import openpyxl
import matplotlib.pyplot as plt
from PySimpleGUI import T, Text, Input, FileBrowse, Button, Checkbox, Window, WIN_CLOSED, Listbox, In, ColorChooserButton, FolderBrowse
import os

# Retourne une liste avec la valeur des cellules d'une plage d'une feuille


def read_range_cells(sheet, r):
    cells = sheet[r[0]:r[1]]
    l = []
    for c in cells:
        if c[0].value == None:
            l.append(0)
        else:
            l.append(c[0].value)
    return l


def find_ranges(sheet, word_topic, word_time):
    I_topic = J_topic = -1
    # On cherche la première cellule des activités ou du temps
    for i in range(1, 101):
        for j in range(1, 101):
            if sheet.cell(i, j).value == word_topic:
                # Cette activité est en fait 2 ligne en dessous du mot clé recherché, on ajoute donc 2...
                first_cell_topic = sheet.cell(i + 2, j).coordinate
                I_topic = i + 2
                J_topic = j
                break
    if I_topic == -1:
        return None
    for i in range(J_topic, 101):
        if sheet.cell(I_topic - 2, i).value == word_time:
            # Cette activité est en fait 2 ligne en dessous du mot clé recherché, on ajoute donc 2...
            first_cell_time = sheet.cell(I_topic, i).coordinate
            J_time = i
            break
    for i in range(I_topic, 101):
        if (sheet.cell(i, J_topic).value) == None:
            last_cell_topic = sheet.cell(i - 1, J_topic).coordinate
            last_cell_time = sheet.cell(i - 1, J_time).coordinate
            break
    return [[first_cell_topic, last_cell_topic], [first_cell_time, last_cell_time]]


def update_dico(sheet, dico, word_time='Total'):
    c = -1
    ranges = find_ranges(sheet, 'HEATMAP', word_time)
    range_topic = ranges[0]
    range_time = ranges[1]
    if range_time == None or range_topic == None:
        print("Sheet " + sheet.title + " does not contain the HEATMAP and Total Keywords")
        return -1
    keys = read_range_cells(sheet, range_topic)
    values = read_range_cells(sheet, range_time)
    # On prend la taille des valeurs du dico -> Nbr de semaines déjà remplies
    M = 0
    if bool(dico):
        M = len(list(dico.values())[0])
    for k in keys:
        c += 1
        if k in dico.keys():
            dico[k].append(values[c])
        else:
            # On met tout plein de 0 puisque c'est une nouvelle entrée et qu'il faut compenser les ratés
            l = M * [0]
            l.append(values[c])
            dico[k] = l
    # Ne pas oublier d'ajouter des 0 pour les activités n'ayant pas eu lieu cette semaine
    for k in set(dico.keys()) - set(keys):
        dico[k].append(0)
    return 0
# ====================================================================================================
#                                                Graph
# ====================================================================================================


def stackplot(dico, Weeks, fntsize, color_map):
    if fntsize == "":
        fntsize = 18
    else:
        fntsize = int(fntsize)
    weeks = []
    for i in Weeks:
        # On ne garde que le numéro de semaine, ça prend trop de place sinon...
        # Normalement il y a un underscore -> on prend tout ce qu'il y a avant.
        # Sinon on ne garde que les 3 premiers caractères
        tmp = i.split("_")[0]
        if len(tmp) > 3:
            tmp = tmp[:3]
        weeks.append(tmp)
    fig, ax = plt.subplots()
    X = range(len(weeks))
    if (len(weeks)) > 30:
        for i in range(0, len(weeks) - 1, 2):
            weeks[i] = ""
    if color_map == "":
        plt.stackplot(X, list(dico.values()), labels=dico.keys())
    else:
        plt.stackplot(X, list(dico.values()), labels=dico.keys(), colors=color_map)
    plt.xticks(X, weeks, fontsize=fntsize)
    plt.yticks(fontsize=fntsize)
    plt.ylabel("Nombre de jours", fontsize=fntsize * 1.25)
    plt.xlabel("Semaines", fontsize=fntsize * 1.25)
    plt.title("Activités de l'équipe Test", fontsize=fntsize * 1.25)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5, fontsize=fntsize)
    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')
    plt.subplots_adjust(bottom=0.2)
    plt.show()
# ====================================================================================================
#                                                Pie
# ====================================================================================================


def pie(dico):
    # L'idée est de sommer chaque temps puis de le diviser par le temps total
    dico_total = {}
    for k in dico.keys():
        dico_total[k] = sum(dico[k])
    temps_total = sum(dico_total.values())
    # Deux boucles quasi identiques à la suite c'est moche, je sais...
    for k in dico_total.keys():
        dico_total[k] /= temps_total
    explode = [0] * len(dico.keys())
    explode[1] = 0.05
    plt.pie(dico_total.values(), explode=explode, labels=dico_total.keys(), autopct='%1.1f%%', startangle=30)
    plt.show()
# ====================================================================================================
#                                                Write
# ====================================================================================================


def write(dico, weeks, path=os.getenv("userprofile") + "\\Downloads\\Output.xlsx"):
    # On écrit le dico dans une feuille excel
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, len(weeks) + 1):
        ws.cell(i + 1, 1, weeks[i - 1])
    col = 1
    for k in dico.keys():
        col += 1
        ws.cell(1, col, k)
        row = 2
        for l in dico[k]:
            ws.cell(row, col, l)
            row += 1
    wb.save(filename=path)
# ====================================================================================================
#                                       Recup d'info dans la GUI
# ====================================================================================================


def interface_input(filepath=""):
    layout = [
        [T("")],
        [Text("Choose a file: "),
         Input(filepath, key="file"),
         FileBrowse()],
        [Text("Choose a tester: "),
         Input('Total', key="tester")],
        [Checkbox('Show Stackplot', default=True, key="stackplot")],
        [Checkbox('Write output in xlsx file', default=True, key="write"),
         Input(os.path.join(os.getenv("userprofile"), "Downloads"), key="folder"),
         FolderBrowse("Browse", initial_folder=os.path.join(os.getenv("userprofile"), "Downloads"))],
        [Checkbox('Show Pie (awful design)', default=False, key="pie")],
        [Button('Run')]]
    # Create the window
    window = Window('TestTeam Activity', layout)
    dico = {}
    while True:
        # Display and interact with the Window
        event, values = window.read()
        if event == WIN_CLOSED:
            break
        if event == "Run":
            dico["file"] = values["file"]
            dico["tester"] = values["tester"]
            dico["stackplot"] = values["stackplot"]
            dico["write"] = values["write"]
            dico["pie"] = values["pie"]
            dico["folder"] = values["folder"]
            break
    window.close()
    return dico
# ====================================================================================================
#                                     GUI de selection de la plage
# ====================================================================================================


def interface_data_range(path, weeks):
    layout = [[T(""), Text(path)],
              [Text("Choose first week"), Listbox(weeks, size=(30, 3), key="first")],
              [Text("Choose last week"), Listbox(weeks, size=(30, 3), key="last")],
              [Button('Run', key="Run")]]
    # Create the window
    window = Window('Data Range', layout)
    # Display and interact with the Window
    event, values = window.read()
    if event == "Run":
        return [window.Element('first').Widget.curselection()[0], window.Element('last').Widget.curselection()[0]]
# ====================================================================================================
#                                       Détection des feuilles Excel
# ====================================================================================================


def filter_sheets(wb_obj):
    L = []
    for sheetname in wb_obj.sheetnames:
        # On considère qu'une feuille est exploitable si elle contient un S au début
        if (sheetname[0] == "S" or sheetname[0] == "s") and any(char.isdigit() for char in sheetname):
            L.append(sheetname)
    return L

# ====================================================================================================
#                                       GUI de paramètrage graphique
# ====================================================================================================


def Color_Choosing_UI(activ_dico, weeks):
    dico = {}
    layout = []
    color_list = ['#1f77b4', '#aec7e8', '#ff7f0e', '#ffbb78', '#2ca02c', '#98df8a', '#d62728', '#ff9896', '#9467bd', '#c5b0d5',
                  '#8c564b', '#c49c94', '#e377c2', '#f7b6d2', '#7f7f7f', '#c7c7c7', '#bcbd22', '#dbdb8d', '#17becf', '#9edae5'] * 10
    cnt = -1
    act_lst = list(activ_dico.keys())
    layout.append([Text("Edit the colors and click on try to see the result")])
    layout.append([Text("Close the window when you're done")])
    for i in act_lst:
        cnt += 1
        colo = color_list[cnt]
        dico[i] = colo
        layout.append(
            [Text(i),
             In("", visible=False, enable_events=True, key='set_line_color_' + i),
             ColorChooserButton(
                 "", size=(1, 1),
                 target="set_line_color_" + i, button_color=(colo, colo),
                 border_width=1, key=i)],)
    layout.append([Text("Font size :"), Input("18", key="font", size=(3, 3))])
    layout.append([Button('Try'), Button('Close')])
    window = Window('Color picking', layout)
    while True:
        # Display and interact with the Window
        event, values = window.read()
        if event == WIN_CLOSED or event == "Close":
            break
        for i in act_lst:
            if event == "set_line_color_" + i:
                window.Element(i).update(button_color=values["set_line_color_" + i])
                dico[i] = values["set_line_color_" + i]
        if event == "Try":
            stackplot(activ_dico, weeks, int(values["font"]), list(dico.values()))
    window.close()
    return dico


# ====================================================================================================
#                                                Main
# ====================================================================================================
UI = interface_input()
xlsx_file = UI["file"]
tester = UI["tester"]
stack = UI["stackplot"]
write_output = UI["write"]
pie_chart = UI["pie"]
folder = UI["folder"]
# ===============================================================================
wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
dico = {}
weeks = []
sheets = filter_sheets(wb_obj)
[first, last] = interface_data_range(xlsx_file, sheets)
for sheetname in sheets[last:first + 1]:
    if update_dico(wb_obj[sheetname], dico, tester) == 0:
        weeks.append(sheetname)
legend = list(dico.keys())
# On se remet dans le bon sens
weeks.reverse()
null_keys = []
for i in dico.keys():
    if all(v == 0 for v in dico[i]):
        null_keys.append(i)
    dico[i].reverse()
# On vire les entrées avec que des 0
for i in null_keys:
    dico.pop(i, None)
if stack:
    Color_Choosing_UI(dico, weeks).values()
if pie_chart:
    pie(dico)
if write_output:
    write(dico, weeks, os.path.join(folder, "output.xlsx"))
